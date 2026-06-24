import openpyxl, io, json, calendar

SRC = r'C:\Users\stoyan.s\Desktop\_Семеен бюджет 2026.xlsx'
OUT = r'C:\Projects\FinApp\tools\FinApp.Seed\family.json'
MONTHS = ['January','February','March','April','May','June']

wb = openpyxl.load_workbook(SRC, data_only=True)

def isnum(v): return isinstance(v, (int, float))

# Top-level category names (codes 1..8) from the `codes` tab.
codes = wb['codes']
cat_names = {}
for r in range(1, codes.max_row+1):
    a, b = codes.cell(r,1).value, codes.cell(r,2).value
    if isnum(a) and 1 <= a <= 8 and float(a).is_integer():
        cat_names[int(a)] = str(b).strip()

def top_level(code):
    code = int(code)
    return code if code < 100 else code // 100

def exp_day(k):
    if k is None: return 1
    if hasattr(k, 'month'):           # Excel mis-parsed dd/mm as US m/d -> user's day = datetime.month
        return k.month
    s = str(k).strip().replace('-', '/').split('/')
    try: return int(s[0])
    except: return 1

# January defines the contributor template (name + category per income row); other months only have amounts.
def income_amounts(ws):
    """Amounts under the 'Приход' header, split from the trailing total via running sum."""
    pr = next((r for r in range(1, 20) if str(ws.cell(r,1).value).strip() == 'Приход'), None)
    if pr is None: return []
    amts, run = [], 0.0
    for r in range(pr+1, pr+12):
        b = ws.cell(r,2).value
        if not isnum(b):
            if amts: break
            continue
        b = float(b)
        if amts and abs(b - run) < 0.01:   # the total row
            break
        amts.append(b); run += b
    return amts

jan = wb['January']
pr = next(r for r in range(1, 20) if str(jan.cell(r,1).value).strip() == 'Приход')
template = []
for i, _ in enumerate(income_amounts(jan)):
    r = pr + 1 + i
    template.append((str(jan.cell(r,1).value).strip(), str(jan.cell(r,3).value).strip()))

months = []
for mi, name in enumerate(MONTHS, start=1):
    ws = wb[name]
    last_day = calendar.monthrange(2026, mi)[1]

    # Opening = the three fund rows at the top (single fund in the app = their sum).
    opening = sum(float(ws.cell(r,2).value) for r in (2,3,4) if isnum(ws.cell(r,2).value))

    contribs = []
    for i, amt in enumerate(income_amounts(ws)):
        who, cat = template[i] if i < len(template) else template[-1]
        contribs.append({'who': who, 'amount': amt, 'cat': cat})

    budgets = {}
    for r in range(2, 10):
        d, f = ws.cell(r,4).value, ws.cell(r,6).value
        if isnum(d) and isnum(f):
            budgets[str(top_level(d))] = float(f)

    savings_saved, savings_initial = [], []
    for r in (25, 26, 27):
        savings_saved.append(float(ws.cell(r,7).value) if isnum(ws.cell(r,7).value) else 0.0)
        savings_initial.append(float(ws.cell(r,6).value) if isnum(ws.cell(r,6).value) else 0.0)

    expenses = []
    for r in range(2, ws.max_row+1):
        j, l = ws.cell(r,10).value, ws.cell(r,12).value
        if isnum(j) and isnum(l) and l > 0:
            if str(int(j)).startswith('99'): continue
            day = min(max(exp_day(ws.cell(r,11).value), 1), last_day)
            expenses.append({'cat': top_level(j), 'amount': round(float(l), 2), 'day': day})

    months.append({
        'name': name, 'index': mi,
        'opening': float(opening) if isnum(opening) else None,
        'contributions': contribs,
        'budgets': budgets,
        'savingsInitial': savings_initial if mi == 1 else None,
        'savingsSaved': savings_saved,
        'expenses': expenses,
    })

bucket_names = [str(wb['January'].cell(r,5).value).strip() for r in (25,26,27)]

data = {
    'account': 'Family',
    'currency': 'EUR',
    'fundName': 'Наличност',
    'categories': cat_names,            # {code: name} for 1..8
    'contribCategories': sorted({c['cat'] for m in months for c in m['contributions']}),
    'contributors': sorted({c['who'] for m in months for c in m['contributions']}),
    'savingBuckets': bucket_names,      # row 25/26/27 -> bucket index 0/1/2
    'months': months,
}
io.open(OUT, 'w', encoding='utf-8').write(json.dumps(data, ensure_ascii=False, indent=1))

# Console-safe summary (no Cyrillic)
print('categories:', sorted(cat_names))
print('contributors:', len(data['contributors']), 'contribCats:', len(data['contribCategories']))
for m in months:
    print(f"{m['name']:8} open={m['opening']} contrib={len(m['contributions'])} "
          f"budgets={len(m['budgets'])} expenses={len(m['expenses'])} "
          f"saved={[round(x,0) for x in m['savingsSaved']]}")
